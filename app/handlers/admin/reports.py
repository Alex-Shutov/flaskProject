import datetime
import re

import pandas as pd
from django.db.models.fields import return_None
from openpyxl import Workbook

from database import get_orders, get_product_info_with_params, get_all_products_with_stock

from database import get_user_info_by_id

from app_types import SaleType

from database import get_product_with_type

from database import get_detailed_orders
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


def sanitize_sheet_name(name):
    # Удаляем или заменяем недопустимые символы в имени листа
    return re.sub(r'[\/\\\?\*\[\]\:]', '_', name)


from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import datetime
import re

def generate_sales_report(start_date, end_date, type_id, filename="sales_report.xlsx"):
    print(type_id)
    orders_data = get_detailed_orders(start_date, end_date,type_id)

    workbook = Workbook()
    sheet_all = workbook.active
    sheet_all.title = "Общий отчет"

    # Заголовки основных колонок
    headers = ["ID заказа", "Продукт", "Свойство продукта", "Тип продукта", "Менеджер", "Курьер", "Упаковщик", "Дата продажи"]

    # Уникальные параметры
    type_columns = set()
    product_columns = set()
    property_columns = set()

    # Определяем уникальные ключи для столбцов
    for order in orders_data:

        # type_columns.update(order.get('type_product_params', {}).keys())
        product_columns.update(order.get('product_values', {}).keys())
        property_columns.update(order.get('product_param_values', {}).keys())

    # Добавляем параметры продуктов и свойств в заголовки
    # headers.extend(list(type_columns))
    headers.extend(list(product_columns))
    headers.extend(list(property_columns))

    # Записываем заголовки
    sheet_all.append(headers)

    # Цвета для разных типов параметров
    type_param_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")  # Голубой для параметров типа
    product_param_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Зеленый для параметров продукта

    # Словарь для отдельных листов по менеджерам
    manager_sheets = {}

    # Заполняем таблицу
    for order in orders_data:
        print(order)
        print('order')
        order_id = order.get('id')
        product_name = order.get('product_name', '-')
        product_param_title = order.get('product_param_title', '-')
        type_name = order.get('type_product', '-')

        # Убираем двойной символ "@" из username
        manager_name = order.get('manager_name', '-').replace('@@', '@')
        courier_name = order.get('courier_name', 'Не указан').replace('@@', '@')
        packer_name = order.get('packer_name', 'Не указан').replace('@@', '@')

        sale_date = order.get('closed_date', '-')
        if sale_date and isinstance(sale_date, datetime.date):
            sale_date = sale_date.strftime('%d.%m.%Y')

        # Основные данные строки
        row = [order_id, product_name, product_param_title, type_name, manager_name, courier_name, packer_name, sale_date]

        # Добавляем значения параметров
        params = {
            **order.get('type_product_params', {}),
            **order.get('product_values', {}),
            **order.get('product_param_values', {})
        }

        # Добавляем значения для всех параметров, заменяя отсутствующие параметры на '-'
        # for col in type_columns:
        #     value = params.get(col, '-')
        #     if isinstance(value, dict):
        #         value = ', '.join(value.keys())
        #     row.append(value)
        for col in product_columns:
            value = params.get(col, '-')
            if isinstance(value, dict):
                value = ', '.join(value.keys())
            row.append(value)
        for col in property_columns:
            value = params.get(col, '-')
            if isinstance(value, dict):
                value = ', '.join(value.keys())
            row.append(value)

        # Добавляем строку в основной лист
        sheet_all.append(row)

        # Создаем листы для менеджеров
        manager = manager_name
        if manager not in manager_sheets:
            sheet_name = re.sub(r'[\/\\\?\*\[\]\:]', '_', manager)
            sheet_manager = workbook.create_sheet(title=sheet_name)
            sheet_manager.append(headers)
            manager_sheets[manager] = sheet_manager
        manager_sheets[manager].append(row)

    # Удаляем дубликаты столбцов
    unique_headers = []
    # deleted_cols = []
    # for col_index, col_name in enumerate(headers):
    #     if col_name not in unique_headers:
    #         unique_headers.append(col_name)
    #     else:
    #         # Удаляем содержимое в колонке с дубликатом, если оно идентично
    #         for row in sheet_all.iter_rows(min_row=2, min_col=col_index+1, max_col=col_index+1):
    #             value = row[0].value
    #             # Проверяем предыдущий столбец на идентичное значение и оставляем только одно
    #             if col_index not in deleted_cols and value == sheet_all.cell(row=row[0].row, column=unique_headers.index(col_name) + 1).value:
    #                 print(123)
    #                 sheet_all.delete_cols(col_index,1)
    #                 deleted_cols.append(col_index)
    #
    #         for sheet in manager_sheets.values():
    #             for row in sheet.iter_rows(min_row=2, min_col=col_index + 1, max_col=col_index + 1):
    #                 value = row[0].value
    #                 if value == sheet_all.cell(row=row[0].row, column=unique_headers.index(col_name) + 1).value:
    #                     sheet_all.delete_rows(col_index, 1)
    #                     row[0].value = None
    col_indexes_to_delete = []

    for col_index, col_name in enumerate(headers):
        if col_name not in unique_headers:
            unique_headers.append(col_name)
        else:
            col_indexes_to_delete.append(col_index + 1)  # +1 так как индексы в openpyxl 1-based

    # Удаляем дубликаты для всех листов, начиная с последнего, чтобы не сдвигать индексы
    for col_index in reversed(col_indexes_to_delete):
        sheet_all.delete_cols(col_index)
        for sheet in manager_sheets.values():
            sheet.delete_cols(col_index)

    # # Применение цветов к столбцам
    # for col_index, col_name in enumerate(headers[8:], start=8):  # начиная с индекса параметров
    #     fill = type_param_fill if col_name in type_columns else product_param_fill
    #     for row in sheet_all.iter_rows(min_row=2, min_col=col_index + 1, max_col=col_index + 1):
    #         for cell in row:
    #             cell.fill = fill
    #
    #     # Применение цветов к менеджерским листам
    #     for sheet in manager_sheets.values():
    #         for row in sheet.iter_rows(min_row=2, min_col=col_index + 1, max_col=col_index + 1):
    #             for cell in row:
    #                 cell.fill = fill

        # Применение цветов к столбцам

    def apply_colors(sheet):
        for col_idx, col_name in enumerate(unique_headers[8:], start=8):  # начиная с индекса параметров
            print(col_idx)
            print(col_name)
            print('col_name')
            # print(type_columns)
            print(property_columns)
            print('property_columns')
            print(product_columns)
            print('property_columns')
            if col_name in product_columns:
                fill = type_param_fill
            elif col_name in property_columns:
                fill = product_param_fill
            else:
                continue  # не окрашиваем колонки, такие как "Дата продажи" и прочие базовые колонки
            for cell in sheet[get_column_letter(col_idx+1)]:
                if cell.row > 1:  # Пропускаем заголовок
                    cell.fill = fill
    apply_colors(sheet_all)
    for sheet in manager_sheets.values():
        apply_colors(sheet)

    # Устанавливаем автоширину для всех столбцов
    for sheet in workbook.worksheets:
        for col in range(1, len(unique_headers) + 1):
            max_length = max(len(str(cell.value)) for cell in sheet[get_column_letter(col)])
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(col)].width = adjusted_width

    # Сохраняем файл
    workbook.save(filename)
    return filename


def generate_stock_report(type_id):
    # Получаем данные о всех типах продуктов и их остатках
    products = get_all_products_with_stock(type_id)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Stock Report"

    # Цвета для разных типов параметров
    type_param_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    product_param_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    price_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    main_product_fill = PatternFill(start_color="DDA0DD", end_color="DDA0DD", fill_type="solid")

    headers = [
        "Продукт",
        "Свойство продукта",
        "Основной продукт",
        "Цена продажи",
        "Цена авито доставки",
        "Остаток"
    ]

    # Создаем множества для хранения уникальных параметров,
    # исключая direct_price и delivery_price из параметров продукта
    product_columns = set()
    property_columns = set()

    # Собираем уникальные ключи параметров из всех продуктов
    for product_type, items in products.items():
        for item in items:
            product_values = item.get('product_values', {}).copy()
            # Удаляем цены из параметров продукта, так как они будут отдельными колонками
            if 'direct_price' in product_values:
                product_values.pop('direct_price')
            if 'delivery_price' in product_values:
                product_values.pop('delivery_price')
            product_columns.update(product_values.keys())
            property_columns.update(item.get('product_param_values', {}).keys())

    # Объединяем все ключи параметров в заголовки
    headers.extend(list(product_columns))
    headers.extend(list(property_columns))
    sheet.append(headers)

    # Заполняем строки данными
    for product_type, items in products.items():
        for item in items:
            product_name = item['name']
            stock = item['stock']
            param_title = item['param_title']
            is_main_product = "Да" if item.get('is_main_product', False) else "Нет"

            # Копируем product_values, чтобы не изменять оригинальный словарь
            product_values = item.get('product_values', {}).copy()
            # Извлекаем цены
            sale_price = product_values.pop('direct_price', '-')
            avito_delivery_price = product_values.pop('delivery_price', '-')

            # Основные данные строки
            row = [
                product_name,
                param_title,
                is_main_product,
                sale_price,
                avito_delivery_price,
                stock
            ]

            # Добавляем оставшиеся параметры продукта
            for col in product_columns:
                value = product_values.get(col, '-')
                row.append(value)

            # Добавляем параметры свойств
            product_param_values = item.get('product_param_values', {})
            for col in property_columns:
                value = product_param_values.get(col, '-')
                row.append(value)

            sheet.append(row)

    # Применение цветов к колонкам
    # Основной продукт
    for cell in sheet[get_column_letter(3)]:
        if cell.row > 1:
            cell.fill = main_product_fill

    # Цены
    for col in range(4, 6):  # Колонки с ценами
        for cell in sheet[get_column_letter(col)]:
            if cell.row > 1:
                cell.fill = price_fill

    # Параметры продукта и свойств
    for col_index, col_name in enumerate(headers[6:], start=6):
        fill = type_param_fill if col_name in product_columns else product_param_fill
        for cell in sheet[get_column_letter(col_index + 1)]:
            if cell.row > 1:
                cell.fill = fill

    # Применяем форматирование
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Устанавливаем автоширину для всех столбцов
    for col in range(1, len(headers) + 1):
        max_length = max(len(str(cell.value)) for cell in sheet[get_column_letter(col)])
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(col)].width = adjusted_width

    # Добавляем границы
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    # Сохраняем файл
    report_path = "stock_report.xlsx"
    workbook.save(report_path)
    return report_path