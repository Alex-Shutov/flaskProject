from decimal import Decimal

from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from app_types import SaleTypeRu, OrderTypeRu
from database import get_connection

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from database import get_setting_value

from app_types import TripStatusRu

from app_types import TrackNumberStatusRu
from database import get_packing_info


def generate_detailed_sales_report(start_date, end_date, filename="detailed_sales_report.xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Полный отчет"

    # Обновленные цвета для статусов
    status_colors = {
        'active': 'A7FC92',  # Зеленый
        'in_packing': 'C1E1C1',  # Салатовый
        'ready_to_delivery': 'FFE599',  # Желтый
        'in_delivery': '9CC3E4',  # Синий
        'refund': 'D3D3D3',  # Серый
        'closed': 'FF9999',  # Красный
        'partly_delivered': 'DEB887'  # Коричневый
    }

    # Создаем легенду на отдельном листе


    # Заголовки
    headers = [
        "ID заказа",
        "Дата продажи",
        "Тип продажи",
        "Продукт",
        "Свойство продукта",
        "Основной продукт",
        "Статус продукта",
        "Доставка/Трекинг",
        "Статус упаковки",  # Новая колонка
        "Причина переупаковки",  # Новая колонка
        "Статус заказа",
        "Менеджер",
        "Показал",
        "Упаковщик",
        "Курьер",
        "Заметка менеджера",
        "Заметка курьера",
        "Оплата менеджеру",
        "Оплата показавшему",
        "Оплата упаковщику",
        "Сумма продажи",
        "Сумма доставки",
        "Итоговая сумма"
    ]

    # Стили
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Применяем заголовки
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Получаем данные о продажах
    sales_data = get_detailed_order_data(start_date, end_date)
    main_products_count = 0
    total_sales_sum = 0
    total_delivery_sum = 0
    total_final_sum = 0
    current_row = 2
    current_order_id = None
    total_viewer_sum = 0
    order_start_row = current_row

    for order in sales_data:
        if current_order_id != order['id']:
            items = order['items']
            total_sale_price = 0
            viewer_payment = 0

            if order.get('viewer_name'):
                # Если есть показывающий, используем разделение цен
                total_sale_price = order['main_products_price']
                viewer_payment = order['additional_products_price']
            else:
                # Если показывающего нет, считаем общую сумму
                total_sale_price = sum(float(item.get('sale_price', 0)) for item in items)


            # if len(items) == 1:
            #     # Для одного продукта - одна строка
            #     single_row = [
            #         order['id'],
            #         order['created_at'].strftime("%d.%m.%Y"),
            #         SaleTypeRu[order['order_type'].upper()].value,
            #         items[0]['product_name'],
            #         items[0]['param_title'],
            #         items[0]['status'],
            #         get_delivery_info(order),
            #         get_delivery_zone(order),  # Новая функция для получения зоны доставки
            #         order['status'],
            #         f"{order['manager_name']} (@{order['manager_username']})",
            #         get_courier_info(order),
            #         order.get('note', ''),  # Заметка менеджера
            #         order.get('delivery_note', ''),  # Заметка курьера
            #         order.get('total_price', 0),
            #         order.get('delivery_price', 0),
            #         calculate_total_sum(order)
            #     ]
            #
            #     row_color = status_colors.get(order['status'], 'FFFFFF')
            #     for col, value in enumerate(single_row, 1):
            #         cell = sheet.cell(row=current_row, column=col)
            #         cell.value = value
            #         cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            #         cell.border = thin_border
            #         cell.alignment = Alignment(vertical='center', wrap_text=True)
            #         # Делаем ID заказа жирным
            #         if col == 1:
            #             cell.font = Font(bold=True)
            #
            #     current_row += 1

            # else:
            # Для нескольких продуктов - группировка
            # Основная строка заказа
            main_products_in_order = sum(1 for item in items if item.get('is_main_product'))
            order_row = [
                order['id'],
                order['created_at'].strftime("%d.%m.%Y"),
                SaleTypeRu[order['order_type'].upper()].value,
                f"{len(items)} продуктов",
                "",
                f"{main_products_in_order} осн.",
                "",
                get_delivery_info(order),
                "",
                "",
                OrderTypeRu[order['status'].upper()].value,
                f"{order['manager_name']} ({order['manager_username']})",
                f"{order['viewer_name']} ({order['viewer_username']})" if order['viewer_name'] else "",
                f"{order['packer_name']} ({order['packer_username']})" if order['packer_name'] else "",
                get_courier_info(order),
                order.get('note', ''),
                order.get('delivery_note', ''),
                total_sale_price,
                viewer_payment,
                calculate_packing_cost(order),
                order.get('total_price', 0),
                order.get('delivery_price', 0),
                calculate_total_sum(order)
            ]

            row_color = status_colors.get(order['status'], 'FFFFFF')
            for col, value in enumerate(order_row, 1):
                cell = sheet.cell(row=current_row, column=col)
                cell.value = value
                cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                # Делаем ID заказа жирным в основной строке
                if col == 1:
                    cell.font = Font(bold=True)

            current_row += 1
            order_start_row = current_row - 1

            # Строки с продуктами
            for item in items:
                packing_status,repacking_reason="",""
                if order['order_type']=='avito':
                    tmp, repacking_reason = get_packing_info(order['id'],item.get('tracking_number'))
                    packing_status = TrackNumberStatusRu[tmp.upper()].value
                product_row = [
                    order['id'],
                    order['created_at'].strftime("%d.%m.%Y"),
                    SaleTypeRu[order['order_type'].upper()].value,
                    item['product_name'],
                    item['param_title'],
                    'Да' if item['is_main_product'] else 'Нет',
                    OrderTypeRu[item['status'].upper()].value,
                    get_delivery_info(order) if order['order_type'] == 'delivery' else item.get('tracking_number', ''),
                    packing_status,
                    repacking_reason,
                    OrderTypeRu[order['status'].upper()].value,  # Дублируем статус заказа
                    f"{order['manager_name']} ({order['manager_username']})",  # Дублируем менеджера
                    f"{order['viewer_name']} ({order['viewer_username']})" if order['viewer_name'] else "",
                    f"{order['packer_name']} ({order['packer_username']})" if order.get('packer_name') else "",
                    # Дублируем упаковщика
                    get_courier_info(order),  # Дублируем курьера
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    ""
                ]
                if item['is_main_product']:
                    main_products_count += 1

                for col, value in enumerate(product_row, 1):
                    cell = sheet.cell(row=current_row, column=col)
                    cell.value = value
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    # ID заказа в строках продуктов НЕ жирный
                    if col == 1:
                        cell.font = Font(bold=False)

                current_row += 1

                # Группируем строки продуктов
                if current_row > order_start_row + 1:
                    sheet.row_dimensions.group(order_start_row + 1, current_row - 1, outline_level=1)

            current_order_id = order['id']


    # Настройка ширины столбцов и форматирование
    for col in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 15

    # Для числовых столбцов устанавливаем формат
    for col in ['K', 'L', 'M']:  # Столбцы с суммами
        for row in range(2, current_row):
            cell = sheet[f"{col}{row}"]
            if cell.value:
                cell.number_format = '#,##0.00'

    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    sheet.freeze_panes = 'A2'

    summary_row = current_row + 2
    sheet.cell(row=summary_row, column=1, value="ИТОГИ:").font = Font(bold=True)

    # Добавляем итоговую информацию
    sheet.cell(row=summary_row, column=3, value=f"Всего основных продуктов: {main_products_count}")
    sheet.cell(row=summary_row + 1, column=3, value=f"Сумма продаж итоговая: {total_sales_sum}")
    sheet.cell(row=summary_row + 2, column=3, value=f"Сумма доставки итоговая: {total_delivery_sum}")
    sheet.cell(row=summary_row + 3, column=3, value=f"Итоговая сумма: {total_final_sum}")

    # Настраиваем фильтры, исключая определенные столбцы
    filter_columns = list(range(1, len(headers) + 1))
    # Исключаем колонки с заметками и суммами
    exclude_columns = [8, 15, 16, 17, 18, 19, 20, 21] # Индексы колонок для исключения из фильтрации
    filter_columns = [col for col in range(1, len(headers) + 1) if col not in exclude_columns]

    # Применяем фильтры только к нужным колонкам
    for col in filter_columns:
        col_letter = get_column_letter(col)
        sheet.auto_filter.add_sort_condition(f"{col_letter}:{col_letter}")

    # Добавляем формулы для автоматического подсчета итогов с учетом фильтров
    # Обновляем формулы
    main_products_formula = (
        f'=SUMPRODUCT(SUBTOTAL(3,OFFSET({get_correct_column_letter("main_products")}2,ROW({get_correct_column_letter("main_products")}2:{get_correct_column_letter("main_products")}{current_row})-ROW({get_correct_column_letter("main_products")}2),0)),'
        f'--({get_correct_column_letter("main_products")}2:{get_correct_column_letter("main_products")}{current_row}="Да"))'
    )
    viewer_payment_formula = f'=SUBTOTAL(9,{get_correct_column_letter("viewer_payment")}2:{get_correct_column_letter("viewer_payment")}{current_row})'
    manager_payment_formula = f'=SUBTOTAL(9,{get_correct_column_letter("manager_payment")}2:{get_correct_column_letter("manager_payment")}{current_row})'
    packer_payment_formula = f'=SUBTOTAL(9,{get_correct_column_letter("packer_payment")}2:{get_correct_column_letter("packer_payment")}{current_row})'
    sales_sum_formula = f'=SUBTOTAL(9,{get_correct_column_letter("sale_sum")}2:{get_correct_column_letter("sale_sum")}{current_row})'
    delivery_sum_formula = f'=SUBTOTAL(9,{get_correct_column_letter("delivery_sum")}2:{get_correct_column_letter("delivery_sum")}{current_row})'
    final_sum_formula = f'=SUBTOTAL(9,{get_correct_column_letter("total_sum")}2:{get_correct_column_letter("total_sum")}{current_row})'

    # Создаем список итоговых строк в нужном порядке
    summary_rows = [
        ("Всего основных продуктов:", main_products_formula),
        ("Сумма оплаты менеджерам:", manager_payment_formula),
        ("Сумма оплаты показавшим:", viewer_payment_formula),
        ("Сумма оплаты упаковщикам:", packer_payment_formula),
        ("Сумма продаж итоговая:", sales_sum_formula),
        ("Сумма доставки итоговая:", delivery_sum_formula),
        ("Итоговая сумма:", final_sum_formula)
    ]

    # Добавляем итоговые строки
    for idx, (label, formula) in enumerate(summary_rows):
        current_summary_row = summary_row + idx

        # Добавляем заголовок
        cell = sheet.cell(row=current_summary_row, column=3)
        cell.value = label
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right', vertical='center')

        # Добавляем значение
        cell = sheet.cell(row=current_summary_row, column=4)
        cell.value = formula
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.number_format = '#,##0.00'

    # Настраиваем ширину колонок
    max_width = max(len(label) for label, _ in summary_rows)
    sheet.column_dimensions[get_column_letter(3)].width = max_width + 2
    sheet.column_dimensions[get_column_letter(4)].width = 15

    workbook = generate_courier_trips_report(workbook, start_date, end_date)

    legend_sheet = workbook.create_sheet("Легенда")
    legend_sheet.cell(row=1, column=1, value="Статус").font = Font(bold=True)
    legend_sheet.cell(row=1, column=2, value="Цвет").font = Font(bold=True)

    for idx, (status, color) in enumerate(status_colors.items(), start=2):
        legend_sheet.cell(row=idx, column=1, value=OrderTypeRu[status.upper()].value)
        cell = legend_sheet.cell(row=idx, column=2)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    legend_sheet.column_dimensions['A'].width = 20
    legend_sheet.column_dimensions['B'].width = 10

    workbook.save(filename)
    return filename


def get_correct_column_letter(column_name):
    """Получаем правильную букву колонки для каждого типа данных"""
    columns = {
        'main_products': 'F',  # Основные продукты
        'manager_payment': 'R',  # Оплата менеджеру
        'viewer_payment': 'S',  # Оплата показавшему
        'packer_payment': 'T',  # Оплата упаковщику
        'sale_sum': 'U',  # Сумма продажи
        'delivery_sum': 'V',  # Сумма доставки
        'total_sum': 'W'  # Итоговая сумма
    }
    return columns.get(column_name)

def get_delivery_info(order):
    """Формирует информацию о доставке"""
    if order['order_type'] == 'delivery':
        return f"Адрес: {order.get('delivery_address', 'Не указан')}"
    return ''
# Добавляем новую вспомогательную функцию
def get_delivery_zone(order):
    """Получает информацию о зоне доставки"""
    if order['order_type'] == 'delivery':
        return order.get('delivery_zone', {}).get('name', 'Не указана')
    return ''


def calculate_packing_cost(order):
    """Рассчитывает стоимость упаковки для заказа"""
    if order['order_type'] != 'avito':
        return 0

    # Получаем базовую цену упаковки
    packing_price = get_setting_value('packing_price')

    # Используем количество упакованных коробок, а не количество трекномеров
    return order.get('packed_boxes_count', 0) * packing_price


def get_courier_info(order):
    """Формирует информацию о курьере"""

    if order.get('courier_name'):
        return f"{order['courier_name']} ({order['courier_username']})"
    return ''


def calculate_total_sum(order):
    total_price = float(order.get('total_price', 0) or 0)  # Если None, то 0
    delivery_price = float(order.get('delivery_price', 0) or 0)  # Если None, то 0
    if order['order_type'] in ['delivery', 'sdek', 'pek', 'luch']:
        return total_price + delivery_price
    return total_price

def get_detailed_order_data(start_date, end_date):
    """
    Получает детальную информацию о заказах за период
    """
    with get_connection() as conn:
        with conn.cursor() as cursor:
            query = """
                WITH order_items_info AS (
                SELECT 
                    oi.order_id,
                    oi.tracking_number,
                    jsonb_agg(
                        jsonb_build_object(
                            'product_name', p.name,
                            'param_title', pp.title,
                            'status', oi.status,
                            'tracking_number', oi.tracking_number,
                            'is_main_product', p.is_main_product,
                             'sale_price', p.sale_price  
                        )
                    ) as items_info,
                    SUM(CASE WHEN p.is_main_product THEN p.sale_price ELSE 0 END) as main_products_price,
                    SUM(CASE WHEN NOT p.is_main_product THEN p.sale_price ELSE 0 END) as additional_products_price
                FROM order_items oi
                JOIN products p ON oi.product_id = p.id
                JOIN product_params pp ON oi.product_param_id = pp.id
                GROUP BY oi.order_id, oi.tracking_number
            ),
            delivery_info AS (
                SELECT 
                    o.id as order_id,
                    ct.total_price as delivery_price,
                    cu.name as courier_name,
                    cu.username as courier_username,
                    COALESCE(dz.name, 'Не указана') as zone_name
                FROM orders o
                LEFT JOIN order_items oi ON o.id = oi.order_id
                LEFT JOIN trip_items ti ON oi.id = ti.order_item_id
                LEFT JOIN courier_trips ct ON ti.trip_id = ct.id
                LEFT JOIN users cu ON o.courier_id = cu.id
                LEFT JOIN delivery_zones dz ON ct.zone_id = dz.id
                GROUP BY o.id, ct.total_price, cu.name, cu.username, dz.name
            )
            SELECT 
                o.id,
                o.created_at,
                o.order_type,
                o.status,
                o.note,
                o.total_price,
                o.delivery_address,
                o.delivery_note,
                o.packed_boxes_count,
                m.name as manager_name,
                m.username as manager_username,
                p.name as packer_name,
                p.username as packer_username,
                v.name as viewer_name,
                v.username as viewer_username,
                tp.title as type_product_name,
                oi.items_info,
                oi.main_products_price,
                oi.additional_products_price,
                di.delivery_price,
                di.courier_name,
                di.courier_username,
                o.delivery_sum,
                di.zone_name
            FROM orders o
            JOIN users m ON o.manager_id = m.id
            LEFT JOIN users p ON o.packer_id = p.id
            LEFT JOIN users v ON o.viewer_id = v.id
            JOIN order_items oit ON o.id = oit.order_id
            JOIN products p_prod ON oit.product_id = p_prod.id
            JOIN type_product tp ON p_prod.type_id = tp.id
            LEFT JOIN order_items_info oi ON o.id = oi.order_id
            LEFT JOIN delivery_info di ON o.id = di.order_id
            WHERE o.created_at BETWEEN %s AND %s
            GROUP BY 
                o.id, o.created_at, o.order_type, o.status, o.note,
                o.total_price, o.delivery_address, o.delivery_note, o.packed_boxes_count,
                m.name, m.username, 
                p.name, p.username,
                v.name, v.username,
                tp.title, oi.items_info,
                oi.main_products_price, oi.additional_products_price,
                di.delivery_price, di.courier_name, di.courier_username,
                o.delivery_sum, di.zone_name
            ORDER BY o.created_at DESC
            """

            cursor.execute(query, (start_date, end_date))
            results = []

            for row in cursor.fetchall():
                order_dict = {
                    'id': row[0],
                    'created_at': row[1],
                    'order_type': row[2],
                    'status': row[3],
                    'note': row[4],
                    'total_price': row[5],
                    'delivery_address': row[6],
                    'delivery_note': row[7],
                    'packed_boxes_count':row[8],
                    'manager_name': row[9],
                    'manager_username': row[10],
                    'packer_name': row[11],
                    'packer_username': row[12],
                    'viewer_name': row[13],
                    'viewer_username': row[14],
                    'type_product_name': row[15],
                    'items': row[16] if isinstance(row[16], list) else [],
                    'main_products_price': row[17] or 0,
                    'additional_products_price': row[18] or 0,
                    'delivery_price': row[19] or row[22],  # Используем delivery_sum если нет delivery_price
                    'courier_name': row[20],
                    'courier_username': row[21]
                }
                results.append(order_dict)

            return results


def get_courier_trips_data(start_date, end_date):
    with get_connection() as conn:
        with conn.cursor() as cursor:
            query = """
                WITH trip_items_info AS (
                    SELECT 
                        ti.trip_id,
                        jsonb_agg(
                            jsonb_build_object(
                                'order_id', o.id,
                                'delivery_address', o.delivery_address,
                                'tracking_number', oi.tracking_number,
                                'status', oi.status,
                                'order_type', o.order_type
                            )
                        ) as items_info,
                        MAX(o.order_type) as order_type -- Получаем тип доставки для поездки
                    FROM trip_items ti
                    JOIN order_items oi ON ti.order_item_id = oi.id
                    JOIN orders o ON oi.order_id = o.id
                    GROUP BY ti.trip_id
                )
                SELECT 
                    ct.id AS trip_id,
                    ct.status AS trip_status,
                    ct.total_price,
                    ct.created_at,
                    ct.completed_at,
                    u.name as courier_name,
                    u.username as courier_username,
                    dz.name as zone_name,
                    dz.base_price,
                    dz.additional_item_price,
                    tii.order_type, -- Тип доставки (Avito или Доставка)
                    tii.items_info,
                    MAX(o.delivery_note) as courier_note
                FROM courier_trips ct
                JOIN users u ON ct.courier_id = u.id
                LEFT JOIN delivery_zones dz ON ct.zone_id = dz.id
                JOIN trip_items_info tii ON ct.id = tii.trip_id
                LEFT JOIN LATERAL (
                    SELECT o.delivery_note
                    FROM orders o
                    WHERE o.id = ANY(ARRAY(
                        SELECT (jsonb_array_elements(tii.items_info)->>'order_id')::int
                    ))
                ) o ON true
                WHERE ct.created_at BETWEEN %s AND %s
                GROUP BY 
                    ct.id, ct.status, ct.total_price, ct.created_at, ct.completed_at,
                    u.name, u.username, dz.name, dz.base_price, dz.additional_item_price, tii.items_info, tii.order_type
                ORDER BY ct.created_at DESC
            """
            cursor.execute(query, (start_date, end_date))
            return cursor.fetchall()




def generate_courier_trips_report(workbook, start_date, end_date):
    trips_sheet = workbook.create_sheet("Поездки")

    # Цвета для статусов поездок
    trip_status_colors = {
        'created': 'A7FC92',  # Зеленый
        'in_progress': '9CC3E4',  # Синий
        'completed': 'FF9999',  # Красный
        'cancelled': 'D3D3D3'  # Серый
    }

    # Заголовки
    headers = [
        "ID поездки",
        "Тип доставки",
        "Курьер",
        "Адрес/Трекномер",
        "Зона доставки",
        "Базовая цена зоны",
        "Доп. цена зоны",
        "Статус товара",
        "Статус поездки",
        "Стоимость поездки",
        "Заметка курьера"
    ]

    # Применяем заголовки и стили
    for col, header in enumerate(headers, 1):
        cell = trips_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    # Получаем данные о поездках
    trips_data = get_courier_trips_data(start_date, end_date)

    current_row = 2
    for trip in trips_data:
        (trip_id, status, total_price, created_at, completed_at, courier_name,
         courier_username, zone_name, base_price, additional_price,
         order_type, items_info, courier_note) = trip

        # Основная строка поездки
        trip_row = [
            trip_id,
            "",
            f"{courier_name} ({courier_username})",
            "",
            "",
            "",
            "",
            "",
            TripStatusRu[status.upper()].value,
            total_price,
            courier_note or ''
        ]

        # Применяем стили и цвета
        row_color = trip_status_colors.get(status, 'FFFFFF')
        for col, value in enumerate(trip_row, 1):
            cell = trips_sheet.cell(row=current_row, column=col)
            cell.value = value
            cell.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if col == 1:
                cell.font = Font(bold=True)
            if isinstance(value, (int, float, Decimal)):
                cell.number_format = '#,##0.00'

        current_row += 1
        items_start_row = current_row

        # Строки с товарами
        for item in items_info:
            item_order_type = item['order_type'].lower()
            item_row = [
                trip_id,
                SaleTypeRu[item_order_type.upper()].value,
                f"{courier_name} ({courier_username})",
                item['delivery_address'] if item_order_type == 'delivery' else item['tracking_number'],
                zone_name if item_order_type == 'delivery' else "",
                base_price if item_order_type == 'delivery' else "",
                additional_price if item_order_type == 'delivery' else "",
                OrderTypeRu[item['status'].upper()].value,
                TripStatusRu[status.upper()].value,
                "",
                ""
            ]

            for col, value in enumerate(item_row, 1):
                cell = trips_sheet.cell(row=current_row, column=col)
                cell.value = value
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                if isinstance(value, (int, float, Decimal)):
                    cell.number_format = '#,##0.00'

            current_row += 1

        # Группируем строки товаров
        if current_row > items_start_row:
            trips_sheet.row_dimensions.group(items_start_row, current_row - 1, outline_level=1)

    summary_row = current_row + 1
    trips_sheet.cell(row=summary_row, column=1, value="ИТОГО:").font = Font(bold=True)
    total_formula = f'=SUBTOTAL(9,J2:J{current_row})'
    cell = trips_sheet.cell(row=summary_row, column=10)  # Колонка "Стоимость поездки"
    cell.value = total_formula
    cell.font = Font(bold=True)
    cell.number_format = '#,##0.00'

    # Настройка ширины колонок
    for col in range(1, len(headers) + 1):
        max_length = 0
        for row in range(1, current_row + 1):
            cell = trips_sheet.cell(row=row, column=col)
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        trips_sheet.column_dimensions[get_column_letter(col)].width = min(adjusted_width, 50)

    filter_columns = [1, 2, 3, 5, 8, 9]  # Индексы нужных колонок для фильтрации

    # Настраиваем автофильтр для всей таблицы
    trips_sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # Добавляем сортировку только для нужных колонок
    for col in filter_columns:
        col_letter = get_column_letter(col)
        trips_sheet.auto_filter.add_sort_condition(f"{col_letter}:{col_letter}")


    return workbook